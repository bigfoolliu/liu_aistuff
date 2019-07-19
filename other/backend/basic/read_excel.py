#!/usr/bin/env python3
# -*- coding:utf-8 -*-


import openpyxl

file_path = "./read_excel_test.xlsx"

# 得到一个WorKBook对象
file_wb = openpyxl.load_workbook(file_path)
print(file_wb, type(file_wb))

# 得到一个表单的列表
sheets = file_wb.sheetnames
print(sheets)

# sheet1 = file_wb.get_sheet_by_name("Sheet1")

# 当前活跃的表单
active_sheet = file_wb.active
print(active_sheet)
print(active_sheet["A1"])
print(active_sheet["A1"].value)

# 获取总的数据行数或列数
rows = active_sheet.max_row
columns = active_sheet.max_column
# print(active_sheet.max_row, active_sheet.max_column)

row_range = active_sheet[1:rows]
column_range = active_sheet["A:B"]

data = []
for row in active_sheet.iter_rows(min_row=2, max_row=rows, max_col=columns):
    # item = {"worker_id": None, "worker_name": None}
    item = {}
    print("----------------")
    item["worker_id"] = row[0].value
    item["name"] = row[1].value
    item["platform"] = row[2].value
    data.append(item)

print(data)
